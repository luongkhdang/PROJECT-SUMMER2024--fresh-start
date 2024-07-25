import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet1 =  wb['Sheet1']

    for row in range(2,sheet1.max_row+1):
        cell = sheet1.cell(row,3)
        corrected_price = cell.value*0.9
        corrected_price_cell = sheet1.cell(row,4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet1,
              min_row = 2,
              max_row = sheet1.max_row,
              min_col=4,
              max_col=4)

    chart1 = BarChart()
    chart1.add_data(values)
    sheet1.add_chart(chart1, 'e2')

    wb.save(filename)